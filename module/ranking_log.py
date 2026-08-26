import pandas as pd
from datetime import datetime
from typing import List, Dict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class Ranking_Log:    
    """
    Ranking log class for recording and saving leaderboard data for uploads, downloads, and deletions.

    Attributes:
        RANKING_COLUMNS (list): Column names for the ranking table.
        logs (List[Dict[str, str]]): List storing ranking entries.
        log_file (str): Save path for the log file.
    """
    RANKING_COLUMNS = ["排名", "使用者", "次數", "檔案大小", "姓名", "電子郵件"]

    def __init__(self):
        """
        Initialize the Ranking_Log instance, setting up the log list and file path.
        """
        self.logs: List[Dict[str, str]] = []
        self.log_file = self.get_log_path()
        self.total_sizes = {
            "upload": 0,
            "download": 0,
            "delete": 0
        }

    def get_log_path(self) -> str:
        """
        Generate the save path for the log file, using the current timestamp in the filename and saving to the Desktop.

        Returns:
            str: Full path of the log file.
        """
        date_str = datetime.now().strftime("%Y-%m-%d-%H_%M_%S")
        desktop_path = Path.home() / "Desktop"
        return str(desktop_path / f"NAS_Ranking_Log_{date_str}.xlsx")

    def add_log(self, ranking_type: str, rank: int, username: str, count: int, size: int, name: str, email: str):
        """
        Add a ranking entry to the log list.

        Parameters:
            ranking_type (str): Type of ranking (e.g., "upload", "download", "delete").
            rank (int): Rank position.
            username (str): Username.
            count (int): Event count.
            size (int): Total file size.
            name (str): User's full name.
            email (str): User's email address.

        Exceptions:
            Exception: Raised with an error message if adding the entry fails.
        """
        try:
            entry = {
                "排名": str(rank),
                "使用者": username,
                "次數": str(count),
                "檔案大小": str(size),
                "姓名": name,
                "電子郵件": email,
                "類型": ranking_type
            }
            self.logs.append(entry)
            self.total_sizes[ranking_type] += size
        except Exception as e:
            raise Exception(f"Failed to add ranking log entry: {str(e)}")
        
    def save_to_excel(self) -> bool:
        """
        Save ranking logs to an Excel file containing three worksheets for upload, download, and delete, applying formatted styles.

        Returns:
            bool: True if saved successfully; False if there are no log entries.

        Exceptions:
            Exception: Raised with an error message if saving fails.
        """
        if not self.logs:
            return False
        try:
            df = pd.DataFrame(self.logs, columns=self.RANKING_COLUMNS + ["類型"])
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            title_font = Font(bold=True, size=14)
            title_alignment = Alignment(horizontal='center', vertical='center')

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color='BFD1E5', end_color='BFD1E5', fill_type='solid')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            data_alignment = Alignment(horizontal='center', vertical='center')
            
            column_widths = {
                'A': 8,   # Rank
                'B': 15,  # Username
                'C': 10,  # Count
                'D': 15,  # File Size
                'E': 15,  # Name
                'F': 25   # Email
            }

            for ranking_type, sheet_name, title in [
                ("upload", "上傳排行榜", "上傳次數排行榜 (總大小: {})".format(self.format_size(self.total_sizes["upload"]))),
                ("download", "下載排行榜", "下載次數排行榜 (總大小: {})".format(self.format_size(self.total_sizes["download"]))),
                ("delete", "刪除排行榜", "刪除次數排行榜 (總大小: {})".format(self.format_size(self.total_sizes["delete"])))
            ]:
                type_df = df[df["類型"] == ranking_type][self.RANKING_COLUMNS]
                
                if type_df.empty:
                    continue

                worksheet = workbook.create_sheet(title=sheet_name)
                worksheet.merge_cells('A1:F1')
                title_cell = worksheet['A1']
                title_cell.value = title
                title_cell.font = title_font
                title_cell.alignment = title_alignment

                for col_num, value in enumerate(self.RANKING_COLUMNS, 1):
                    cell = worksheet.cell(row=2, column=col_num)
                    cell.value = value
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill
                    cell.border = border

                for row_idx, row_data in enumerate(type_df.itertuples(), 3):
                    for col_idx, value in enumerate(row_data[1:], 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = self.format_size(value) if col_idx == 4 else value  # Format size column
                        cell.alignment = data_alignment
                        cell.border = border

                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
            
            workbook.save(self.log_file)
            self.logs.clear()
            return True
        except Exception as e:
            raise Exception(f"Failed to save ranking log: {str(e)}")

    def format_size(self, size):
        """
        Format byte count into a human-readable file size string (KB, MB, GB).
        
        Parameters:
            size (int/str): File size in bytes.
            
        Returns:
            str: Formatted file size string.
        """
        try:
            size = int(size)
            if size == 0:
                return "0 KB"
            for unit in ['bytes', 'KB', 'MB', 'GB']:
                if size < 1024.0:
                    return f"{size:.2f} {unit}"
                size /= 1024.0
            return f"{size:.2f} GB"
        except:
            return str(size)